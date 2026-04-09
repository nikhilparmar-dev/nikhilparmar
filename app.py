from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from supabase import create_client, Client
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
from datetime import date
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import io

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'rajivnagar-shala-secret-2025')

# supabase connection
url = os.getenv('SUPABASE_URL')
key = os.getenv('SUPABASE_KEY')
supabase: Client = create_client(url, key)

# -----------------------------------------------
# HELPER FUNCTIONS
# -----------------------------------------------

def is_logged_in():
    return 'user_id' in session

def is_teacher():
    return session.get('role') == 'teacher'

def is_principal():
    return session.get('role') == 'principal'

def get_grade(total, out_of=60):
    pct = (total / out_of) * 100
    if pct >= 90: return 'A+'
    elif pct >= 75: return 'A'
    elif pct >= 60: return 'B'
    elif pct >= 45: return 'C'
    else: return 'D'

# -----------------------------------------------
# GENERAL ROUTES
# -----------------------------------------------

@app.route('/')
def index():
    if not is_logged_in():
        return redirect(url_for('login'))
    if is_principal():
        return redirect(url_for('principal_dashboard'))
    return redirect(url_for('teacher_dashboard'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        # TEMP ADMIN LOGIN (bypass)
if username == "admin" and password == "admin":
    session['user_id'] = "admin"
    session['name'] = "Admin"
    session['username'] = "admin"
    session['role'] = "principal"
    session['class'] = ""

    return redirect(url_for('principal_dashboard'))

        # find user in supabase
        result = supabase.table('users').select('*').eq('username', username).execute()

        if result.data:
            user = result.data[0]
            if check_password_hash(user['password_hash'], password):
                # set session
                session['user_id'] = user['id']
                session['name'] = user['name']
                session['username'] = user['username']
                session['role'] = user['role']
                session['class'] = user.get('class_assigned', '')

                if user['role'] == 'principal':
                    return redirect(url_for('principal_dashboard'))
                return redirect(url_for('teacher_dashboard'))
            else:
                error = 'Wrong password. Please try again.'
        else:
            error = 'Username not found. Please check.'

    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['POST'])
def change_password():
    if not is_logged_in():
        return redirect(url_for('login'))

    old_pass = request.form.get('old_password')
    new_pass = request.form.get('new_password')

    # verify old password
    user = supabase.table('users').select('*').eq('id', session['user_id']).execute()
    if user.data:
        if check_password_hash(user.data[0]['password_hash'], old_pass):
            supabase.table('users').update({
                'password_hash': generate_password_hash(new_pass)
            }).eq('id', session['user_id']).execute()
            return jsonify({'success': True, 'msg': 'Password changed!'})
        else:
            return jsonify({'success': False, 'msg': 'Old password is wrong.'})

    return jsonify({'success': False, 'msg': 'Something went wrong.'})


# -----------------------------------------------
# TEACHER ROUTES
# -----------------------------------------------

@app.route('/teacher/dashboard')
def teacher_dashboard():
    if not is_logged_in() or not is_teacher():
        return redirect(url_for('login'))

    my_class = session['class']
    today = str(date.today())

    # get students count
    students = supabase.table('students').select('id').eq('class', my_class).execute()
    total = len(students.data)

    # today attendance
    att = supabase.table('attendance').select('*').eq('date', today).execute()
    student_ids = [s['id'] for s in students.data]
    my_att = [a for a in att.data if a['student_id'] in student_ids]

    present = len([a for a in my_att if a['status'] == 'P'])
    absent = len([a for a in my_att if a['status'] == 'A'])
    late = len([a for a in my_att if a['status'] == 'L'])

    # low attendance students
    all_students = supabase.table('students').select('*').eq('class', my_class).order('roll_no').execute()
    all_att = supabase.table('attendance').select('*').execute()

    low_att_students = []
    for s in all_students.data:
        s_att = [a for a in all_att.data if a['student_id'] == s['id']]
        if len(s_att) > 0:
            pct = round((len([a for a in s_att if a['status'] == 'P']) / len(s_att)) * 100)
            if pct < 75:
                low_att_students.append({'name': s['name'], 'roll': s['roll_no'], 'pct': pct})

    return render_template('teacher/dashboard.html',
        total=total,
        present=present,
        absent=absent,
        late=late,
        low_count=len(low_att_students),
        low_students=low_att_students[:5],
        today=today
    )


@app.route('/teacher/attendance')
def teacher_attendance():
    if not is_logged_in() or not is_teacher():
        return redirect(url_for('login'))

    my_class = session['class']
    today = str(date.today())

    students = supabase.table('students').select('*').eq('class', my_class).order('roll_no').execute()

    # check if today attendance already marked
    att_today = supabase.table('attendance').select('*').eq('date', today).execute()
    student_ids = [s['id'] for s in students.data]
    att_map = {}
    for a in att_today.data:
        if a['student_id'] in student_ids:
            att_map[a['student_id']] = a['status']

    # calculate overall attendance % for each student
    all_att = supabase.table('attendance').select('*').execute()
    for s in students.data:
        s_att = [a for a in all_att.data if a['student_id'] == s['id']]
        if len(s_att) > 0:
            s['att_pct'] = round((len([a for a in s_att if a['status'] == 'P']) / len(s_att)) * 100)
        else:
            s['att_pct'] = 100

    return render_template('teacher/attendance.html',
        students=students.data,
        att_map=att_map,
        today=today
    )


@app.route('/teacher/attendance/save', methods=['POST'])
def save_attendance():
    if not is_logged_in() or not is_teacher():
        return jsonify({'error': 'Not authorized'}), 401

    data = request.json
    today = str(date.today())

    for student_id, status in data.get('attendance', {}).items():
        # check if already exists for today
        existing = supabase.table('attendance').select('id') \
            .eq('student_id', student_id).eq('date', today).execute()

        if existing.data:
            # update existing record
            supabase.table('attendance').update({'status': status}) \
                .eq('student_id', student_id).eq('date', today).execute()
        else:
            # insert new record
            supabase.table('attendance').insert({
                'student_id': student_id,
                'date': today,
                'status': status,
                'marked_by': session['user_id']
            }).execute()

    return jsonify({'success': True})


@app.route('/teacher/students')
def teacher_students():
    if not is_logged_in() or not is_teacher():
        return redirect(url_for('login'))

    my_class = session['class']
    students = supabase.table('students').select('*').eq('class', my_class).order('roll_no').execute()

    # calculate attendance % for each
    all_att = supabase.table('attendance').select('*').execute()
    for s in students.data:
        s_att = [a for a in all_att.data if a['student_id'] == s['id']]
        if len(s_att) > 0:
            s['att_pct'] = round((len([a for a in s_att if a['status'] == 'P']) / len(s_att)) * 100)
        else:
            s['att_pct'] = 100

    return render_template('teacher/students.html', students=students.data)


@app.route('/teacher/students/add', methods=['POST'])
def add_student():
    if not is_logged_in() or not is_teacher():
        return redirect(url_for('login'))

    data = {
        'name': request.form.get('name', '').strip(),
        'roll_no': request.form.get('roll_no', '').strip(),
        'gr_number': request.form.get('gr_number', '').strip(),
        'class': session['class'],
        'dob': request.form.get('dob', ''),
        'gender': request.form.get('gender', ''),
        'caste': request.form.get('caste', ''),
        'parent_contact': request.form.get('parent_contact', ''),
        'aadhaar_number': request.form.get('aadhaar_number', ''),
        'bank_account': request.form.get('bank_account', ''),
    }

    # upload aadhaar document
    if 'aadhaar_doc' in request.files:
        f = request.files['aadhaar_doc']
        if f and f.filename:
            file_bytes = f.read()
            path = f"docs/{session['class']}/aadhaar_{data['roll_no']}_{f.filename}"
            supabase.storage.from_('school-docs').upload(path, file_bytes,
                {'content-type': f.content_type})
            data['aadhaar_doc_url'] = supabase.storage.from_('school-docs').get_public_url(path)

    # upload ration card document
    if 'ration_doc' in request.files:
        f = request.files['ration_doc']
        if f and f.filename:
            file_bytes = f.read()
            path = f"docs/{session['class']}/ration_{data['roll_no']}_{f.filename}"
            supabase.storage.from_('school-docs').upload(path, file_bytes,
                {'content-type': f.content_type})
            data['ration_doc_url'] = supabase.storage.from_('school-docs').get_public_url(path)

    supabase.table('students').insert(data).execute()
    return redirect(url_for('teacher_students'))


@app.route('/teacher/students/edit', methods=['POST'])
def edit_student():
    if not is_logged_in() or not is_teacher():
        return redirect(url_for('login'))

    student_id = request.form.get('student_id')
    data = {
        'name': request.form.get('name', '').strip(),
        'roll_no': request.form.get('roll_no', '').strip(),
        'gr_number': request.form.get('gr_number', '').strip(),
        'dob': request.form.get('dob', ''),
        'gender': request.form.get('gender', ''),
        'caste': request.form.get('caste', ''),
        'parent_contact': request.form.get('parent_contact', ''),
        'aadhaar_number': request.form.get('aadhaar_number', ''),
        'bank_account': request.form.get('bank_account', ''),
    }

    supabase.table('students').update(data).eq('id', student_id).execute()
    return redirect(url_for('teacher_students'))


@app.route('/teacher/marks')
def teacher_marks():
    if not is_logged_in() or not is_teacher():
        return redirect(url_for('login'))

    my_class = session['class']
    subject = request.args.get('subject', 'GUJ')
    semester = request.args.get('semester', '1')

    students = supabase.table('students').select('*').eq('class', my_class).order('roll_no').execute()

    # get existing marks for this subject and semester
    marks_result = supabase.table('marks').select('*') \
        .eq('subject', subject).eq('semester', semester).execute()
    marks_map = {m['student_id']: m for m in marks_result.data}

    return render_template('teacher/marks.html',
        students=students.data,
        marks_map=marks_map,
        subject=subject,
        semester=semester
    )


@app.route('/teacher/marks/save', methods=['POST'])
def save_marks():
    if not is_logged_in() or not is_teacher():
        return jsonify({'error': 'Not authorized'}), 401

    data = request.json
    subject = data.get('subject')
    semester = data.get('semester')

    for student_id, marks in data.get('marks', {}).items():
        record = {
            'student_id': student_id,
            'subject': subject,
            'semester': semester,
            'written_marks': int(marks.get('written', 0)),
            'participation_marks': int(marks.get('participation', 0)),
            'academic_year': '2025-26',
            'entered_by': session['user_id']
        }

        existing = supabase.table('marks').select('id') \
            .eq('student_id', student_id).eq('subject', subject).eq('semester', semester).execute()

        if existing.data:
            supabase.table('marks').update(record).eq('id', existing.data[0]['id']).execute()
        else:
            supabase.table('marks').insert(record).execute()

    return jsonify({'success': True})


@app.route('/teacher/download')
def teacher_download():
    if not is_logged_in() or not is_teacher():
        return redirect(url_for('login'))
    return render_template('teacher/download.html')


@app.route('/teacher/download/gunslip')
def download_gunslip():
    if not is_logged_in() or not is_teacher():
        return redirect(url_for('login'))

    my_class = session['class']
    semester = request.args.get('semester', '1')

    students = supabase.table('students').select('*').eq('class', my_class).order('roll_no').execute()
    subjects = ['GUJ', 'ENG', 'HINDI', 'SAN', 'MATHS', 'SCI', 'SS']

    # get all marks
    marks_all = supabase.table('marks').select('*').eq('semester', semester).execute()
    marks_by_student = {}
    for m in marks_all.data:
        sid = m['student_id']
        if sid not in marks_by_student:
            marks_by_student[sid] = {}
        marks_by_student[sid][m['subject']] = m

    wb = Workbook()

    # helper for border
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # DATA SHEET
    ws_data = wb.active
    ws_data.title = 'DATA'

    ws_data['C2'] = 'સમિતિ :'
    ws_data['D2'] = 'જિલ્લા શિક્ષણ સમિતિ, મહેસાણા'
    ws_data['B3'] = 'શાળાનું પુરુ નામ :'
    ws_data['D3'] = 'રાજીવનગર પ્રાથમિક શાળા'
    ws_data['B4'] = 'તાલુકો :'
    ws_data['D4'] = 'કડી'
    ws_data['B5'] = 'વર્ષ :'
    ws_data['D5'] = '2025-26'
    ws_data['B6'] = 'ઘોરણ :'
    ws_data['D6'] = my_class.replace('Class ', '')
    ws_data['B7'] = 'રજીસ્ટર સંખ્યા :'
    ws_data['D7'] = len(students.data)
    ws_data['B8'] = '૫રીક્ષા :'
    ws_data['D8'] = 'પ્રથમ સત્રાંત ૫રીક્ષા' if semester == '1' else 'દ્વિતીય સત્રાંત ૫રીક્ષા'
    ws_data['C9'] = 'ક્રમ'
    ws_data['D9'] = 'વિઘાર્થીનું નામ'

    for i, s in enumerate(students.data):
        row = 10 + i
        ws_data.cell(row=row, column=3, value=i + 1)
        ws_data.cell(row=row, column=4, value=s['name'])

    # SUBJECT SHEETS
    subject_names_gu = {
        'GUJ': 'ગુજરાતી',
        'ENG': 'અંગ્રેજી',
        'HINDI': 'હિન્દી',
        'SAN': 'સંસ્કૃત',
        'MATHS': 'ગણિત',
        'SCI': 'વિજ્ઞાન અને ટેક.',
        'SS': 'સામાજીક વિજ્ઞાન'
    }

    for subj in subjects:
        ws = wb.create_sheet(title=subj)

        ws['B3'] = 'રાજીવનગર પ્રાથમિક શાળા'
        ws['B4'] = f'{'પ્રથમ' if semester == '1' else 'દ્વિતીય'} સત્રાંત ૫રીક્ષા : 2025-26'
        ws['B5'] = f'શાળાનું નામ : રાજીવનગર પ્રાથમિક શાળા  તા. : કડી'
        ws['B6'] = f'ઘોરણ : {my_class.replace("Class ", "")}   રજીસ્ટર સંખ્યા : {len(students.data)}'
        ws['B8'] = 'ક્રમ'
        ws['C8'] = 'વિઘાર્થીનું નામ'
        ws['D8'] = f'વિષય :- {subject_names_gu.get(subj, subj)}'

        # question columns header
        for q in range(1, 15):
            ws.cell(row=9, column=3 + q, value=q)
        ws.cell(row=9, column=18, value='કુલ')
        ws.cell(row=9, column=19, value='40 માંથી')

        # student data
        for i, s in enumerate(students.data):
            row = 10 + i
            ws.cell(row=row, column=2, value=i + 1)
            ws.cell(row=row, column=3, value=s['name'])

            # marks
            sid = s['id']
            if sid in marks_by_student and subj in marks_by_student[sid]:
                m = marks_by_student[sid][subj]
                written = m.get('written_marks', 0)
                ws.cell(row=row, column=18, value=written * 2)
                ws.cell(row=row, column=19, value=written)

    # 20 GUN SHEET - class participation
    ws_20 = wb.create_sheet(title='20 GUN')
    ws_20['B3'] = 'રાજીવનગર પ્રાથમિક શાળા'
    ws_20['B4'] = f'{'પ્રથમ' if semester == "1" else "દ્વિતીય"} સત્રાંત ૫રીક્ષા : 2025-26    ઘોરણ : {my_class.replace("Class ", "")}'
    ws_20['B6'] = 'વર્ગખંડમાં વિઘાર્થીની સહભાગિતા આઘારે મૂલ્યાંકન ૫ત્રક'
    ws_20['B8'] = 'ક્રમ'
    ws_20['C8'] = 'વિઘાર્થીનું નામ'

    subj_gu_short = ['ગુજ.', 'અં.', 'હિ.', 'સં.', 'ગ.', 'વિ.', 'સા.']
    for idx, sg in enumerate(subj_gu_short):
        ws_20.cell(row=8, column=4 + idx, value=sg)

    for i, s in enumerate(students.data):
        row = 9 + i
        ws_20.cell(row=row, column=2, value=i + 1)
        ws_20.cell(row=row, column=3, value=s['name'])
        sid = s['id']
        for j, subj in enumerate(subjects):
            if sid in marks_by_student and subj in marks_by_student[sid]:
                p = marks_by_student[sid][subj].get('participation_marks', 0)
                ws_20.cell(row=row, column=4 + j, value=p)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'GUN_SLIP_{my_class}_Sem{semester}.xlsx'
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename)


@app.route('/teacher/download/pdf/<student_id>')
def download_student_pdf(student_id):
    if not is_logged_in():
        return redirect(url_for('login'))

    student = supabase.table('students').select('*').eq('id', student_id).execute()
    if not student.data:
        return 'Student not found', 404

    s = student.data[0]

    # get marks
    marks = supabase.table('marks').select('*').eq('student_id', student_id).execute()
    marks_map = {}
    for m in marks.data:
        key = f"{m['subject']}_sem{m['semester']}"
        marks_map[key] = m

    # get attendance
    att = supabase.table('attendance').select('*').eq('student_id', student_id).execute()
    total_days = len(att.data)
    present_days = len([a for a in att.data if a['status'] == 'P'])
    att_pct = round((present_days / total_days * 100)) if total_days > 0 else 0

    # create PDF
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4

    # header
    c.setFillColor(colors.HexColor('#1a4f8a'))
    c.rect(0, h - 80, w, 80, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 18)
    c.drawString(40, h - 40, 'Rajivnagar Primary School')
    c.setFont('Helvetica', 12)
    c.drawString(40, h - 60, 'Student Profile Report')

    # student name
    c.setFillColor(colors.HexColor('#1e293b'))
    c.setFont('Helvetica-Bold', 16)
    c.drawString(40, h - 110, s['name'])
    c.setFont('Helvetica', 12)
    c.setFillColor(colors.HexColor('#64748b'))
    c.drawString(40, h - 130, f"Class: {s['class']}   Roll No: {s['roll_no']}   GR: {s.get('gr_number', '-')}")

    # divider line
    c.setStrokeColor(colors.HexColor('#e2e8f0'))
    c.line(40, h - 145, w - 40, h - 145)

    # student details
    c.setFillColor(colors.HexColor('#1e293b'))
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, h - 170, 'Student Details')

    details = [
        ('Date of Birth', s.get('dob', '-')),
        ('Gender', s.get('gender', '-')),
        ('Caste', s.get('caste', '-')),
        ('Parent Contact', s.get('parent_contact', '-')),
        ('Aadhaar No.', s.get('aadhaar_number', '-')),
        ('Bank Account', s.get('bank_account', '-')),
    ]

    c.setFont('Helvetica', 11)
    y = h - 195
    for label, value in details:
        c.setFillColor(colors.HexColor('#64748b'))
        c.drawString(40, y, f'{label}:')
        c.setFillColor(colors.HexColor('#1e293b'))
        c.drawString(180, y, str(value))
        y -= 22

    # attendance
    c.line(40, y - 5, w - 40, y - 5)
    y -= 25
    c.setFont('Helvetica-Bold', 12)
    c.setFillColor(colors.HexColor('#1e293b'))
    c.drawString(40, y, 'Attendance')
    y -= 22
    c.setFont('Helvetica', 11)
    c.setFillColor(colors.HexColor('#64748b'))
    c.drawString(40, y, f'Total Days: {total_days}')
    c.drawString(180, y, f'Present: {present_days}')
    c.drawString(300, y, f'Percentage: {att_pct}%')
    y -= 30

    # marks table
    c.line(40, y - 5, w - 40, y - 5)
    y -= 25
    c.setFont('Helvetica-Bold', 12)
    c.setFillColor(colors.HexColor('#1e293b'))
    c.drawString(40, y, 'Marks Summary')
    y -= 22

    subjects = ['GUJ', 'ENG', 'HINDI', 'SAN', 'MATHS', 'SCI', 'SS']
    subj_names = {
        'GUJ': 'Gujarati', 'ENG': 'English', 'HINDI': 'Hindi',
        'SAN': 'Sanskrit', 'MATHS': 'Maths', 'SCI': 'Science', 'SS': 'Social Sci'
    }

    # table header
    c.setFillColor(colors.HexColor('#1a4f8a'))
    c.rect(40, y - 5, w - 80, 20, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 10)
    c.drawString(45, y + 2, 'Subject')
    c.drawString(180, y + 2, 'Sem 1 Written')
    c.drawString(310, y + 2, 'Sem 1 Part.')
    c.drawString(430, y + 2, 'Sem 2 Written')
    y -= 22

    for subj in subjects:
        c.setFillColor(colors.HexColor('#f8fafc'))
        c.rect(40, y - 5, w - 80, 20, fill=True, stroke=False)
        c.setStrokeColor(colors.HexColor('#e2e8f0'))
        c.rect(40, y - 5, w - 80, 20, fill=False, stroke=True)
        c.setFillColor(colors.HexColor('#1e293b'))
        c.setFont('Helvetica', 10)
        c.drawString(45, y + 2, subj_names.get(subj, subj))

        sem1 = marks_map.get(f'{subj}_sem1', {})
        sem2 = marks_map.get(f'{subj}_sem2', {})

        c.drawString(180, y + 2, str(sem1.get('written_marks', '-')))
        c.drawString(310, y + 2, str(sem1.get('participation_marks', '-')))
        c.drawString(430, y + 2, str(sem2.get('written_marks', '-')))
        y -= 22

    c.setFont('Helvetica', 9)
    c.setFillColor(colors.HexColor('#94a3b8'))
    c.drawString(40, 30, 'Rajivnagar Primary School, Rajivnagar, Kadi, Mehsana')
    c.drawString(w - 150, 30, f'Generated on {str(date.today())}')

    c.save()
    buffer.seek(0)

    return send_file(buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'Student_{s["name"]}_Profile.pdf')


# -----------------------------------------------
# PRINCIPAL ROUTES
# -----------------------------------------------

@app.route('/principal/dashboard')
def principal_dashboard():
    if not is_logged_in() or not is_principal():
        return redirect(url_for('login'))

    today = str(date.today())

    all_students = supabase.table('students').select('*').execute()
    all_teachers = supabase.table('users').select('*').eq('role', 'teacher').execute()
    all_att_today = supabase.table('attendance').select('*').eq('date', today).execute()

    total_students = len(all_students.data)
    total_teachers = len(all_teachers.data)
    total_present = len([a for a in all_att_today.data if a['status'] == 'P'])

    # class wise summary
    teachers = all_teachers.data
    teacher_map = {t['class_assigned']: t for t in teachers}

    class_summary = []
    for num in range(1, 9):
        cls = f'Class {num}'
        cls_students = [s for s in all_students.data if s['class'] == cls]
        if cls_students:
            ids = [s['id'] for s in cls_students]
            present = len([a for a in all_att_today.data if a['student_id'] in ids and a['status'] == 'P'])
            att_marked = any(a['student_id'] in ids for a in all_att_today.data)
            t = teacher_map.get(cls, {})
            class_summary.append({
                'class': cls,
                'teacher': t.get('name', 'Not Assigned'),
                'total': len(cls_students),
                'present': present,
                'att_marked': att_marked
            })

    return render_template('principal/dashboard.html',
        total_students=total_students,
        total_teachers=total_teachers,
        total_present=total_present,
        class_summary=class_summary
    )


@app.route('/principal/teachers')
def principal_teachers():
    if not is_logged_in() or not is_principal():
        return redirect(url_for('login'))

    teachers = supabase.table('users').select('*').eq('role', 'teacher').execute()
    return render_template('principal/teachers.html', teachers=teachers.data)


@app.route('/principal/teachers/add', methods=['POST'])
def add_teacher():
    if not is_logged_in() or not is_principal():
        return redirect(url_for('login'))

    supabase.table('users').insert({
        'name': request.form['name'],
        'username': request.form['username'],
        'password_hash': generate_password_hash(request.form['password']),
        'role': 'teacher',
        'class_assigned': request.form['class_assigned']
    }).execute()

    return redirect(url_for('principal_teachers'))


@app.route('/principal/teachers/reset', methods=['POST'])
def reset_teacher_password():
    if not is_logged_in() or not is_principal():
        return redirect(url_for('login'))

    teacher_id = request.form['teacher_id']
    new_password = request.form['new_password']

    supabase.table('users').update({
        'password_hash': generate_password_hash(new_password)
    }).eq('id', teacher_id).execute()

    return redirect(url_for('principal_teachers'))


@app.route('/principal/teachers/delete', methods=['POST'])
def delete_teacher():
    if not is_logged_in() or not is_principal():
        return redirect(url_for('login'))

    teacher_id = request.form['teacher_id']
    supabase.table('users').delete().eq('id', teacher_id).execute()
    return redirect(url_for('principal_teachers'))


if __name__ == '__main__':
    app.run(debug=True)
